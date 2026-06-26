import pytest
from openpyxl import Workbook, load_workbook

from app.core.context import AgentContext, set_agent_context


pytestmark = pytest.mark.no_infrastructure


@pytest.fixture
def excel_context(tmp_path, monkeypatch):
    from app.services.ai.tools import document_paths, generated_file_service

    uploads = tmp_path / "uploads"
    uploads.mkdir()
    source = uploads / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Month", "Amount"])
    sheet.append(["Jan", 10])
    workbook.save(source)
    monkeypatch.setattr(document_paths, "get_data_base_dir", lambda: str(tmp_path))
    async def workspace_root():
        return str(tmp_path / "agent_workspaces")
    monkeypatch.setattr(document_paths, "resolve_workspace_root", workspace_root)
    monkeypatch.setattr(generated_file_service, "generated_files_root", lambda: tmp_path / "generated")
    set_agent_context(AgentContext(
        agent_id="agent", agent_name="Agent", user_id=1, conversation_id="conv",
        authorized_attachment_paths=[str(source)],
    ))
    return source


@pytest.mark.asyncio
async def test_excel_read_range_returns_matrix(excel_context):
    from app.services.ai.tools.excel_document_tool import excel_document_read

    result = await excel_document_read.ainvoke({
        "action": "read_range", "path": str(excel_context),
        "sheet_name": "Sales", "cell_range": "A1:B2",
    })

    assert result["data"]["values"] == [["Month", "Amount"], ["Jan", 10]]


@pytest.mark.asyncio
async def test_excel_write_cells_creates_downloadable_copy(excel_context):
    from app.services.ai.tools.excel_document_tool import excel_document_write

    result = await excel_document_write.ainvoke({
        "action": "write_cells", "path": str(excel_context), "sheet_name": "Sales",
        "cells": [{"address": "B2", "value": 42}], "output_filename": "sales_updated.xlsx",
    })

    assert result["changes"]["written_cells"] == 1
    assert result["artifact"]["download_url"]
    assert load_workbook(excel_context)["Sales"]["B2"].value == 10
